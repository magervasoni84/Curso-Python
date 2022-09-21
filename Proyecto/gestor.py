import tkinter

#Se realiza la importacionde los modulos
import openpyxl
from openpyxl import Workbook
#Usar \\ o / para la ruta
book = openpyxl.load_workbook("C:/Programacion/Curso Python/Proyecto/usuarios.xlsx")

#Era para activar la hoja de trabajo????
hoja1 = book.active



#Min y Max Columnas y Filas Activas
min_col  = book.active.min_column
max_col  = book.active.max_column
min_fila = book.active.min_row
max_fila = book.active.max_row


def saveData():
    dni = cajaDni.get() #Tomo el valor de la caja y lo asigno a dni
    print("Prueba ", hoja1.tables.items())
    print("Valor de Dni", dni) #Toma el Valor
    for valor in hoja1.iter_rows(min_fila, max_col, max_fila, values_only=True):
        print(valor)
        if valor == hoja1.tables['DNI']:
            print("Encontre el DNI")
            #hoja1[CELL.dni] = cajaEdad
    #Poner la ruta completa, sino explota
    #book.save("C:\\Programacion\\Curso Python\\Proyecto\\usuarios.xlsx")

window = tkinter.Tk()
window.title("Gestor")
#tamaño original de la ventana en px
window.config(width=400, height=300)


#Esto es un label DNI
etiqueta = tkinter.Label(window, text="DNI", bg= "blue")
etiqueta.place(x= 40, y= 20)

cajaDni = tkinter.Entry(window)
cajaDni.place(x= 80, y=20)


#Esto es un label Edad
etiqueta = tkinter.Label(window, text="Edad", bg= "blue")
etiqueta.place(x= 40, y= 50)

cajaEdad = tkinter.Entry(window)
cajaEdad.place(x= 80, y=50)







#creacion de un boton Grabar
boton = tkinter.Button(text="Grabar Datos", command= saveData)
boton.place(x= 300, y= 20)

window.mainloop()
